import os
import re
import requests
import fitz  # PyMuPDF for PDF
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract

# ── Set tesseract path (Windows) ──────────────────────────────
import platform
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# On Linux (Streamlit Cloud) tesseract is installed via packages.txt


# ==============================================================
# TEXT EXTRACTION FROM FILES
# ==============================================================

def extract_from_pdf(file) -> str:
    """Extract all text from a PDF file."""
    try:
        doc = fitz.open(stream=file.read(), filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text()
        return text.strip()
    except Exception as e:
        return f"[ERROR reading PDF: {str(e)}]"


def extract_from_docx(file) -> str:
    """Extract all text from a Word document."""
    try:
        doc = Document(file)
        text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        return text.strip()
    except Exception as e:
        return f"[ERROR reading DOCX: {str(e)}]"


def extract_from_excel(file) -> str:
    """Extract all text/data from an Excel file."""
    try:
        wb = load_workbook(file, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"\n[Sheet: {sheet}]\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text += row_text + "\n"
        return text.strip()
    except Exception as e:
        return f"[ERROR reading Excel: {str(e)}]"


def extract_from_image(file) -> str:
    """Use OCR to extract text from an image."""
    try:
        image = Image.open(file)
        text = pytesseract.image_to_string(image)
        return text.strip()
    except Exception as e:
        return f"[ERROR reading image: {str(e)}]"


# ==============================================================
# ELEMENT DETECTION
# ==============================================================

def detect_emojis(text: str) -> list:
    """Find all emojis in the text."""
    emoji_pattern = re.compile(
        "["
        u"\U0001F600-\U0001F64F"
        u"\U0001F300-\U0001F5FF"
        u"\U0001F680-\U0001F9FF"
        u"\U00002702-\U000027B0"
        "]+", flags=re.UNICODE
    )
    return emoji_pattern.findall(text)


def extract_links(text: str) -> list:
    """Find all URLs in the text."""
    url_pattern = re.compile(
        r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|'
        r'(?:%[0-9a-fA-F][0-9a-fA-F]))+'
    )
    return url_pattern.findall(text)


def count_words(text: str) -> int:
    """Count total words in text."""
    return len(text.strip().split())


# ==============================================================
# HYPERLINK VERIFICATION
# ==============================================================

def verify_links(links: list) -> list:
    """
    Check each link to see if it works.
    Returns a list of dicts with url, status, and working flag.
    """
    results = []
    for url in links:
        try:
            response = requests.head(url, timeout=5, allow_redirects=True)
            working = response.status_code < 400
            results.append({
                "url": url,
                "status_code": response.status_code,
                "working": working
            })
        except Exception as e:
            results.append({
                "url": url,
                "status_code": "Error",
                "working": False
            })
    return results

def extract_qr_links(file) -> list:
    """
    Scans an image or GIF for QR codes and extracts the URLs inside them.
    Returns a list of URLs found in QR codes.
    """
    try:
        from pyzbar.pyzbar import decode
        import numpy as np
        import cv2

        image = Image.open(file).convert("RGB")
        img_array = np.array(image)
        img_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)

        decoded_objects = decode(img_bgr)
        qr_links = []
        for obj in decoded_objects:
            data = obj.data.decode("utf-8")
            qr_links.append(data)

        return qr_links
    except Exception as e:
        return []